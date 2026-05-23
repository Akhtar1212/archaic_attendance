import os
import time
import threading
import shutil
import sqlite3
import calendar
import zipfile
import io
from datetime import datetime, timedelta
from statistics import mean

from flask import (
    Flask, render_template, request, redirect, url_for, session,
    jsonify, Response, send_file, send_from_directory
)
import cv2
import numpy as np

APP_DIR = os.path.abspath(os.path.dirname(__file__))
os.makedirs(os.path.join(APP_DIR, "datasets"), exist_ok=True)
os.makedirs(os.path.join(APP_DIR, "proofs"), exist_ok=True)
os.makedirs(os.path.join(APP_DIR, "exports"), exist_ok=True)
os.makedirs(os.path.join(APP_DIR, "backups"), exist_ok=True)

PROOFS_DIR = os.path.join(APP_DIR, "proofs")

# ===== DB & Face Engine =====
from models.db import (
    DB_PATH,
    init_db, add_employee, list_employees, generate_emp_id,
    today_row_for, upsert_checkin, upsert_checkout,
    month_recap_with_proof, rekap_by_employee, list_attendance_by_employee,
    mark_resign, list_today_dashboard,
    submit_leave, list_leave_by_month, approve_leave, reject_leave,
    leaderboard_month,
    set_credentials, get_pegawai_by_credentials
)
from models.face import FaceEngine

# ===== Optional export libs =====
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:
    Workbook = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
except Exception:
    canvas = None

app = Flask(__name__)
app.secret_key = os.environ.get("ARCHAIC_SECRET_KEY", "super-secret-archaic")

ADMIN_USER = os.environ.get("ARCHAIC_ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("ARCHAIC_ADMIN_PASS", "admin123")

# ====== Config dinamis (bisa diubah dari admin) ======
# (Kiosk PIN DIHAPUS – hanya tersisa require_liveness)
CONFIG = {
    "require_liveness": True,
}

# ====== Init ======
init_db()
engine = FaceEngine(
    model_path=os.path.join(APP_DIR, "datasets", "lbph_model.xml"),
    dataset_dir=os.path.join(APP_DIR, "datasets")
)

# ========= Aturan waktu =========
def is_weekend(dt: datetime):
    # Sesuai requirement: Minggu dihitung weekday, weekend hanya Sabtu
    return dt.weekday() == 5  # 5 = Sabtu

def get_operational_datetime(now: datetime):
    """Menangani shift malam: jam 00:00-04:00 dianggap hari operasional kemarin"""
    if now.hour < 4:
        op_date = now - timedelta(days=1)
        return op_date, True # Sedang dalam masa lembur/pulang telat
    return now, False

def get_shift_info(target_date: datetime):
    """Disesuaikan agar menerima target_date (tanggal operasional)"""
    masuk = target_date.replace(hour=11, minute=0, second=0, microsecond=0)
    grace = masuk + timedelta(minutes=10) 
    if is_weekend(target_date):
        # Shift Sabtu pulang Minggu jam 00.00
        end = (target_date + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        end = target_date.replace(hour=23, minute=0, second=0, microsecond=0)
    early = end - timedelta(minutes=15)
    return masuk, grace, end, early

def classify_checkin_status(now: datetime):
    _, grace, _, _ = get_shift_info(now)
    return "Tepat Waktu" if now <= grace else "Terlambat"

def can_checkout(now: datetime):
    _, _, _, early = get_shift_info(now)
    return now >= early

def parse_hms(hms: str):
    try:
        h, m, s = map(int, (hms or "00:00:00").split(":"))
        return h, m, s
    except Exception:
        return 0, 0, 0

def to_seconds(hms: str):
    h, m, s = parse_hms(hms)
    return h * 3600 + m * 60 + s

def secs_to_hms(secs: int):
    if secs is None:
        return "-"
    if secs < 0:
        secs = 0
    h = secs // 3600
    m = (secs % 3600) // 60
    s = secs % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

# ========= Kamera Streamer =========
class CameraStreamer:
    def __init__(self):
        self.cap = None
        self.lock = threading.Lock()
        self.running = False
        self.last_jpeg = None
        self.thread = None
        self.index = 0
        self.backend_name = "INIT"

    def _open(self, index: int):
        trials = [
            (cv2.CAP_MSMF, "MSMF"),
            (cv2.CAP_DSHOW, "DSHOW"),
            (0, "ANY")
        ]
        for backend, name in trials:
            cap = cv2.VideoCapture(int(index), backend)
            if not cap.isOpened():
                cap.release()
                continue
            ok, _ = cap.read()
            if ok:
                self.backend_name = name
                return cap
            cap.release()
        return None

    def start(self, index: int = 0):
        if self.running:
            return True, f"Live sudah berjalan (idx {self.index}, backend {self.backend_name})"
        cap = self._open(index)
        if cap is None:
            return False, f"Gagal membuka kamera index {index}"
        self.cap = cap
        self.index = index
        self.running = True
        self.thread = threading.Thread(target=self._loop, daemon=True)
        self.thread.start()
        return True, f"Live dimulai (idx {index}, backend {self.backend_name})"

    def _loop(self):
        try:
            self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        except Exception:
            pass
        
        while self.running:
            # VALIDASI: Cek apakah kamera masih terhubung
            if self.cap is None or not self.cap.isOpened():
                self.running = False
                break

            ok, frame = self.cap.read()
            if not ok:
                # Jika gagal baca frame, tunggu sebentar lalu coba lagi
                time.sleep(0.1)
                continue
            try:
                cv2.putText(
                    frame,
                    f"Cam {self.index} - {self.backend_name}",
                    (10, 24),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (20, 200, 255),
                    2,
                    cv2.LINE_AA
                )
            except Exception:
                pass
            ret, jpeg = cv2.imencode(".jpg", frame)
            if ret:
                with self.lock:
                    self.last_jpeg = jpeg.tobytes()
            time.sleep(0.03)
        try:
            self.cap.release()
        except Exception:
            pass
        self.cap = None

    def stop(self):
        self.running = False
        return True

    def get_frame(self):
        with self.lock:
            return self.last_jpeg

    def scan(self, max_idx=6):
        found = []
        for idx in range(max_idx):
            for backend, name in [
                (cv2.CAP_MSMF, "MSMF"),
                (cv2.CAP_DSHOW, "DSHOW"),
                (0, "ANY"),
            ]:
                cap = cv2.VideoCapture(idx, backend)
                if not cap.isOpened():
                    cap.release()
                    continue
                ok, frame = cap.read()
                if ok:
                    h, w = frame.shape[:2]
                    found.append({"index": idx, "backend": name, "size": [w, h]})
                    cap.release()
                    break
                cap.release()
        return found

streamer = CameraStreamer()

# ========= Haar untuk liveness =========
FACE_CASCADE = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)
EYE_CASCADE = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_eye.xml"
)

def _decode_jpeg(jpeg_bytes):
    nparr = np.frombuffer(jpeg_bytes, np.uint8)
    return cv2.imdecode(nparr, cv2.IMREAD_COLOR)

# --- REVISI BAGIAN 1: GANTI FUNGSI LIVENESS (STRICT BLINK) ---
def _liveness_check(duration_sec=5.0, step=0.05):
    """
    LOGIKA BLINK (VERSI LONGGAR):
    - minNeighbors diturunkan jadi 3 agar mata lebih mudah terdeteksi.
    - Durasi diperlama jadi 5 detik agar user punya waktu pas-in posisi.
    """
    start = time.time()
    
    blink_stage = 0 
    blink_frames = 0
    
    while time.time() - start < duration_sec:
        frame_jpg = streamer.get_frame()
        if not frame_jpg:
            time.sleep(step)
            continue

        frame = _decode_jpeg(frame_jpg)
        if frame is None:
            time.sleep(step)
            continue

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        gray = cv2.equalizeHist(gray) # Pencerah
        
        faces = FACE_CASCADE.detectMultiScale(gray, 1.1, 4, minSize=(60, 60))
        
        if len(faces) == 0:
            blink_stage = 0
            blink_frames = 0
            time.sleep(step)
            continue

        x, y, w, h = sorted(faces, key=lambda a: a[2] * a[3], reverse=True)[0]
        
        # Fokus area mata
        face_upper = gray[y:y + int(h/1.8), x:x+w]
        
        # TUNING: minNeighbors=3 (Lebih mudah deteksi mata, walau agak noise)
        eyes = EYE_CASCADE.detectMultiScale(face_upper, 1.1, 3, minSize=(10, 10))
        
        is_eye_open = len(eyes) >= 1

        # LOGIKA STATE MACHINE
        if blink_stage == 0:
            # Stage 0: Cari mata terbuka
            if is_eye_open:
                blink_stage = 1
                # print("DEBUG: Mata Terbuka")

        elif blink_stage == 1:
            # Stage 1: Tunggu mata hilang (kedip)
            if not is_eye_open:
                blink_frames += 1
                if blink_frames >= 1: # 1 frame hilang cukup
                    blink_stage = 2
                    # print("DEBUG: Mata Hilang (Kedip)")
            else:
                blink_frames = 0

        elif blink_stage == 2:
            # Stage 2: Tunggu mata balik lagi
            if is_eye_open:
                return True, "blink_verified"

        time.sleep(step)

    return False, "timeout"

# ========= Routes dasar & auth =========
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/admin", methods=["GET", "POST"])
def admin_page():
    if request.method == "POST":
        u = request.form.get("username", "")
        p = request.form.get("password", "")
        if u == ADMIN_USER and p == ADMIN_PASS:
            session["admin"] = True
            return redirect(url_for("admin_page"))
        return render_template("login.html", error="Username / password salah")
    if not session.get("admin"):
        return render_template("login.html")
    return render_template("admin.html", month=datetime.now().strftime("%Y-%m"))

@app.route("/pegawai")
def pegawai_page():
    return render_template("pegawai.html", month=datetime.now().strftime("%Y-%m"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))

# ========= Konfigurasi (Admin) =========
@app.route("/config/get")
def cfg_get():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    return jsonify(CONFIG)

@app.route("/config/set", methods=["POST"])
def cfg_set():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    data = request.json or {}
    if "require_liveness" in data:
        CONFIG["require_liveness"] = bool(data["require_liveness"])
    # kiosk_pin DIHAPUS – tidak diproses lagi
    return jsonify({"ok": True, "config": CONFIG})

# ========= Auth Pegawai =========
@app.route("/auth/pegawai/login", methods=["POST"])
def emp_login():
    data = request.json or {}
    row = get_pegawai_by_credentials(
        (data.get("username") or "").strip(),
        (data.get("password") or "").strip()
    )
    if not row:
        return jsonify({"ok": False, "error": "Username/password salah"}), 401
    pid, nama, jabatan, status = row
    if status != "Aktif":
        return jsonify({"ok": False, "error": "Akun pegawai nonaktif"}), 403
    session["emp_id"], session["emp_name"], session["emp_role"] = pid, nama, jabatan
    return jsonify({"ok": True, "id": pid, "nama": nama, "jabatan": jabatan})

@app.route("/auth/pegawai/logout", methods=["POST"])
def emp_logout():
    session.pop("emp_id", None)
    session.pop("emp_name", None)
    session.pop("emp_role", None)
    return jsonify({"ok": True})

@app.route("/me")
def me():
    return jsonify({
        "logged_in": bool(session.get("emp_id")),
        "id": session.get("emp_id"),
        "nama": session.get("emp_name"),
        "jabatan": session.get("emp_role")
    })

# ========= Kamera (Admin) =========
@app.route("/camera/scan")
def camera_scan():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    return jsonify({"devices": streamer.scan()})

@app.route("/camera/start", methods=["POST"])
def camera_start():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    idx = int((request.json or {}).get("index", 0))
    ok, msg = streamer.start(idx)
    return jsonify({"ok": ok, "message": msg})

@app.route("/camera/stop", methods=["POST"])
def camera_stop():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    streamer.stop()
    return jsonify({"ok": True, "message": "Live dihentikan"})

# ========= Kamera (Pegawai – TANPA PIN) =========
@app.route("/camera/start_public", methods=["POST"])
def camera_start_public():
    # Wajib login sebagai pegawai, tapi tidak perlu PIN lagi
    if not session.get("emp_id"):
        return jsonify({"ok": False, "error": "Harus login sebagai pegawai"}), 403
    idx = int((request.json or {}).get("index", 0))
    if idx < 0 or idx > 5:
        idx = 0
    ok, msg = streamer.start(idx)
    return jsonify({"ok": ok, "message": msg})

# ========= REVISI BAGIAN 2: GANTI FUNGSI VISUAL (THRESHOLD & WARNA) =========

@app.route("/camera/live")
def camera_live():
    if not streamer.running:
        streamer.start(0)

    def generate():
        while True:
            frame_bytes = streamer.get_frame()
            if not frame_bytes:
                time.sleep(0.03)
                continue

            nparr = np.frombuffer(frame_bytes, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

            if frame is None:
                continue

            try:
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                gray = cv2.equalizeHist(gray) # Pencerah otomatis
                
                # Deteksi Wajah
                faces = FACE_CASCADE.detectMultiScale(gray, 1.1, 4, minSize=(60, 60))

                for (x, y, w, h) in faces:
                    # 1. Gambar Kotak Wajah (Default Merah)
                    color = (0, 0, 255) 
                    text_label = "Tidak Dikenali"

                    # Cek Pengenalan (LBPH)
                    try:
                        roi_gray = gray[y:y+h, x:x+w]
                        id_pred, conf = engine.recognizer.predict(roi_gray)

                        if conf < 80: # Threshold 80 (Longgar)
                            color = (0, 255, 0) # Hijau
                            text_label = f"Terdaftar ({int(100 - conf)}%)"
                        else:
                            color = (0, 0, 255)
                            text_label = "Tidak Dikenali"
                    except:
                        pass

                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.rectangle(frame, (x, y-30), (x+w, y), color, cv2.FILLED)
                    cv2.putText(frame, text_label, (x+5, y-5), 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

                    # --- [BARU] VISUALISASI MATA (DEBUGGING) ---
                    # Gambar kotak BIRU di mata agar user tau matanya terdeteksi/tidak
                    face_upper = gray[y:y + int(h/1.8), x:x+w]
                    # minNeighbors=3 agar lebih mudah deteksi mata
                    eyes = EYE_CASCADE.detectMultiScale(face_upper, 1.1, 3, minSize=(10, 10))
                    
                    for (ex, ey, ew, eh) in eyes:
                        # Koordinat mata relatif thd wajah, jadi harus ditambah x, y
                        cv2.rectangle(frame, (x + ex, y + ey), (x + ex + ew, y + ey + eh), (255, 255, 0), 1)
                    # -------------------------------------------

            except Exception as e:
                print(f"[Overlay Error] {e}")

            ret, buffer = cv2.imencode('.jpg', frame)
            if ret:
                frame_final = buffer.tobytes()
                yield (
                    b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                    + frame_final + b"\r\n"
                )
            
            time.sleep(0.03)

    return Response(
        generate(),
        mimetype="multipart/x-mixed-replace; boundary=frame"
    )

# ========= Enroll / Train =========
@app.route("/employees")
def api_employees():
    return jsonify(list_employees())

@app.route("/pegawai/resign", methods=["POST"])
def api_resign():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    data = request.json or {}
    idp = (data.get("id_pegawai") or "").strip()
    if not idp:
        return jsonify({"error": "id_pegawai wajib"}), 400
    try:
        mark_resign(idp)
        return jsonify({"ok": True, "message": "Pegawai di-resign & ID di-recycle"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/enroll/start", methods=["POST"])
def api_enroll_start():
    # Hanya admin yang boleh enroll pegawai baru
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403

    data = request.json or {}
    nama = (data.get("nama") or "").strip()
    jabatan = (data.get("jabatan") or "Kasir").strip()
    count = int(data.get("count") or 40)

    if not nama:
        return jsonify({"error": "Nama wajib diisi"}), 400

    # Generate ID pegawai & simpan ke DB
    emp_id = generate_emp_id(jabatan)
    add_employee(emp_id, nama, jabatan)

    # Capture dataset wajah (dengan quality gate di FaceEngine)
    ok, msg = engine.capture_dataset(
        emp_id,
        nama,
        count=count,
        stream_getter=streamer.get_frame
    )

    if not ok:
        return jsonify({
            "ok": False,
            "message": msg,
            "id": emp_id
        })

    # AUTO TRAIN setelah capture sukses
    trained_count = engine.train()

    return jsonify({
        "ok": True,
        "message": f"{msg} | Model dilatih ulang untuk {trained_count} pegawai.",
        "id": emp_id,
        "trained": trained_count
    })

@app.route("/train")
def api_train():
    n = engine.train()
    return jsonify({"message": f"Training selesai untuk {n} pegawai!", "ok": True})

# ========= Liveness =========
@app.route("/liveness/blink_or_move")
def liveness():
    if not streamer.running:
        return jsonify({"ok": False, "error": "Kamera belum jalan"}), 400
    ok, how = _liveness_check(duration_sec=2.0, step=0.12) # Parameter default diubah di fungsi
    return jsonify({"ok": ok, "method": how or "-"})

# ========= Presensi (recognize) + cooldown =========
_LAST_ACTION = {}

def _cooldown(emp_id: str, sec: int = 30):
    now = time.time()
    last = _LAST_ACTION.get(emp_id, 0)
    if now - last < sec:
        return False, int(sec - (now - last))
    _LAST_ACTION[emp_id] = now
    return True, 0

# --- REVISI BAGIAN 3: GANTI ROUTE ABSEN (STRICT LOGIC) ---
@app.route("/recognize/once")
def api_recognize_once():
    # ========================================================
    # TAHAP 0: LOGIKA WAKTU OPERASIONAL (TAMBAHAN)
    # ========================================================
    now = datetime.now()
    # Menentukan apakah sekarang masih dianggap 'hari kemarin' (sebelum jam 4 pagi)
    op_now, is_late_night = get_operational_datetime(now) # <--- REVISI
    op_date_str = op_now.strftime("%Y-%m-%d")            # <--- REVISI

    # ========================================================
    # TAHAP 1: TANTANGAN WAJIB KEDIP (Liveness Check)
    # ========================================================
    if CONFIG.get("require_liveness", True):
        ok_lv, method = _liveness_check(duration_sec=4.0)
        
        if not ok_lv:
            return jsonify({
                "recognized": False,
                "message": "Gagal Verifikasi: Pastikan Anda BERKEDIP dengan jelas di depan kamera."
            }), 403

    # ========================================================
    # TAHAP 2: PROSES PENGENALAN WAJAH (Setelah Lolos Kedip)
    # ========================================================
    frame = streamer.get_frame()
    if not frame:
        return jsonify({"error": "Kamera error"}), 400

    label, name, conf, face_img = engine.recognize_jpeg(frame)
    
    if not label:
        return jsonify({"recognized": False, "message": "Wajah tidak terdeteksi"})

    if conf > 80:
        return jsonify({
            "recognized": False,
            "message": "Wajah Tidak Dikenali. Pastikan posisi wajah tegak dan pencahayaan cukup."
        }), 403

    # ========================================================
    # TAHAP 3: REKAM ABSENSI (Database)
    # ========================================================
    emp_id = label
    
    ok_cd, remain = _cooldown(emp_id, 30)
    if not ok_cd:
        return jsonify({
            "recognized": True,
            "id": emp_id,
            "nama": name,
            "action": "cooldown",
            "message": f"Tunggu {remain} detik lagi."
        }), 429

    # Mencari data berdasarkan TANGGAL OPERASIONAL
    row = today_row_for(emp_id, op_date_str) # <--- REVISI (Tambah op_date_str)

    if not row or not row["check_in"]:
        # --- Masuk ---
        if is_late_night: # Jika sudah dini hari, dilarang Check-In baru
             return jsonify({"recognized": True, "message": "Sudah lewat batas jam masuk."}), 403
             
        status = classify_checkin_status(now)
        img_path = engine.save_proof(face_img, emp_id, "in", now)
        upsert_checkin(emp_id, now.strftime("%H:%M:%S"), status, img_path, op_date_str) # <--- REVISI
        
        return jsonify({
            "recognized": True,
            "id": emp_id,
            "nama": name,
            "action": "check_in",
            "status": status,
            "message": f"Halo {name}, Berhasil Masuk!"
        })
    else:
        # --- Pulang ---
        # Cek apakah sudah boleh pulang berdasarkan shift tanggal operasional
        _, _, _, early_out = get_shift_info(op_now) # <--- REVISI
        if now < early_out:
            return jsonify({
                "recognized": True, 
                "message": f"Belum jam pulang! (Min jam {early_out.strftime('%H:%M')})"
            })
        
        if row["check_out"]:
            return jsonify({
                "recognized": True,
                "message": "Sudah absen pulang hari ini."
            })
            
        img_path = engine.save_proof(face_img, emp_id, "out", now)
        upsert_checkout(emp_id, now.strftime("%H:%M:%S"), img_path, op_date_str) # <--- REVISI
        
        return jsonify({
            "recognized": True,
            "id": emp_id,
            "nama": name,
            "action": "check_out",
            "status": "Pulang",
            "message": f"Sampai jumpa {name}, Berhasil Pulang!"
        })

# ========= Rekap & Stats =========
@app.route("/rekap_month_with_proof")
def api_rekap_month():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    return jsonify(month_recap_with_proof(month))

@app.route("/rekap_by_employee")
def api_rekap_by_employee():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    emp_id = (request.args.get("id") or "").strip()
    if not emp_id:
        return jsonify({"error": "id pegawai kosong"}), 400
    return jsonify(rekap_by_employee(month, emp_id))

@app.route("/my_rekap")
def api_my_rekap():
    if not session.get("emp_id"):
        return jsonify({"error": "Belum login pegawai"}), 403
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    return jsonify(list_attendance_by_employee(month, session["emp_id"]))

@app.route("/today_board")
def api_today_board():
    return jsonify(list_today_dashboard())

def _late_minutes_for(date_str: str, check_in_hms: str):
    try:
        if not check_in_hms:
            return 0
        h, m, s = map(int, check_in_hms.split(":"))
        cin_sec = h * 3600 + m * 60 + s
        grace_sec = 11 * 3600 + 10 * 60  # 11:10
        return max(0, (cin_sec - grace_sec) // 60)
    except Exception:
        return 0

def _work_secs(check_in_hms: str, check_out_hms: str):
    if not check_in_hms or not check_out_hms:
        return 0
    return max(0, to_seconds(check_out_hms) - to_seconds(check_in_hms))

def _avg_time_str(times):
    times = [t for t in times if t]
    if not times:
        return "-"
    secs = [to_seconds(t) for t in times]
    return secs_to_hms(int(mean(secs)))

def compute_stats(rows):
    present = ontime = late = total_late_mins = total_work_secs = 0
    checkins, checkouts = [], []
    for r in rows:
        # r = [id, id_pegawai, nama, jabatan, tgl, cin, cout, status, bukti_in, bukti_out]
        _, _, _, _, tgl, cin, cout, st, *_ = r
        if cin or cout:
            present += 1
        if st == "Tepat Waktu":
            ontime += 1
        if st == "Terlambat":
            late += 1
            total_late_mins += _late_minutes_for(tgl, cin)
        total_work_secs += _work_secs(cin, cout)
        if cin:
            checkins.append(cin)
        if cout:
            checkouts.append(cout)
    return {
        "present_days": present,
        "on_time": ontime,
        "late": late,
        "late_minutes": int(total_late_mins),
        "work_seconds": int(total_work_secs),
        "work_hms": secs_to_hms(int(total_work_secs)),
        "avg_checkin": _avg_time_str(checkins),
        "avg_checkout": _avg_time_str(checkouts),
    }

@app.route("/stats/summary")
def stats_summary():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    rows = month_recap_with_proof(month)
    return jsonify(compute_stats(rows))

@app.route("/stats/employee")
def stats_employee():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    emp_id = (request.args.get("id") or "").strip()
    if not emp_id:
        return jsonify({"error": "id pegawai kosong"}), 400
    rows = rekap_by_employee(month, emp_id)
    out = compute_stats(rows)
    if rows:
        out["id_pegawai"] = rows[0][1]
        out["nama"] = rows[0][2] or "-"
        out["jabatan"] = rows[0][3] or "-"
    else:
        out["id_pegawai"] = emp_id
    return jsonify(out)

@app.route("/stats/my")
def stats_my():
    if not session.get("emp_id"):
        return jsonify({"error": "Belum login pegawai"}), 403
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    recs = list_attendance_by_employee(month, session["emp_id"])
    rows = []
    for (tgl, cin, cout, st) in recs:
        rows.append([
            None,
            session["emp_id"],
            session.get("emp_name"),
            session.get("emp_role"),
            tgl,
            cin,
            cout,
            st,
            None,
            None,
        ])
    return jsonify(compute_stats(rows))

# ========= Detail Hari & Tren =========
@app.route("/attendance/day_detail")
def day_detail():
    emp_id = (request.args.get("id") or "").strip()
    tgl = (request.args.get("date") or "").strip()
    if not emp_id or not tgl:
        return jsonify({"error": "id dan date wajib"}), 400
    conn = sqlite3.connect(DB_PATH, timeout=10)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT e.id_pegawai, e.nama, e.jabatan, a.tanggal,
               a.check_in, a.check_out, a.status, a.bukti_in, a.bukti_out
        FROM attendance a
        LEFT JOIN employees e ON e.id_pegawai = a.id_pegawai
        WHERE a.id_pegawai = ? AND a.tanggal = ?
        """,
        (emp_id, tgl),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Data tidak ditemukan"}), 404
    idp, nama, jab, tanggal, cin, cout, st, bi, bo = row
    late_m = _late_minutes_for(tanggal, cin)
    dur = secs_to_hms(_work_secs(cin, cout))
    return jsonify({
        "id_pegawai": idp,
        "nama": nama,
        "jabatan": jab,
        "tanggal": tanggal,
        "check_in": cin,
        "check_out": cout,
        "status": st,
        "bukti_in": bi,
        "bukti_out": bo,
        "late_minutes": late_m,
        "work_hms": dur,
    })

@app.route("/trend/month_summary")
def month_trend():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    conn = sqlite3.connect(DB_PATH, timeout=10)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT tanggal,
               SUM(CASE WHEN status = 'Tepat Waktu' THEN 1 ELSE 0 END) AS tepat,
               SUM(CASE WHEN status = 'Terlambat' THEN 1 ELSE 0 END) AS telat
        FROM attendance
        WHERE tanggal LIKE ?
        GROUP BY tanggal
        ORDER BY tanggal
        """,
        (f"{month}%",),
    )
    rows = cur.fetchall()
    conn.close()
    y, m = map(int, month.split("-"))
    days = calendar.monthrange(y, m)[1]
    idx = {r[0]: (r[1], r[2]) for r in rows}
    out = []
    for d in range(1, days + 1):
        t = f"{y:04d}-{m:02d}-{d:02d}"
        tepat, telat = idx.get(t, (0, 0))
        out.append({
            "date": t,
            "ontime": int(tepat),
            "late": int(telat),
        })
    return jsonify(out)

@app.route("/trend/employee_days")
def emp_trend_days():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    emp_id = (request.args.get("id") or "").strip()
    if not emp_id:
        return jsonify({"error": "id kosong"}), 400
    conn = sqlite3.connect(DB_PATH, timeout=10)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT tanggal, check_in, status
        FROM attendance
        WHERE id_pegawai = ? AND tanggal LIKE ?
        ORDER BY tanggal
        """,
        (emp_id, f"{month}%"),
    )
    rows = cur.fetchall()
    conn.close()
    y, m = map(int, month.split("-"))
    days = calendar.monthrange(y, m)[1]
    by_date = {r[0]: r for r in rows}
    out = []
    for d in range(1, days + 1):
        t = f"{y:04d}-{m:02d}-{d:02d}"
        rec = by_date.get(t)
        if rec:
            late = _late_minutes_for(t, rec[1])
            present = 1
        else:
            late = 0
            present = 0
        out.append({
            "date": t,
            "late": int(late),
            "present": present,
        })
    return jsonify(out)

# ========= Export =========
@app.route("/export_csv")
def export_csv():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    rows = month_recap_with_proof(month)
    path = os.path.join(APP_DIR, "exports", f"rekap_{month}.csv")
    with open(path, "w", encoding="utf-8") as f:
        f.write("id,id_pegawai,nama,jabatan,tanggal,masuk,pulang,status,bukti_in,bukti_out\n")
        for r in rows:
            f.write(",".join([str(x or "") for x in r]) + "\n")
    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype="text/csv",
    )

@app.route("/export_excel")
def export_excel():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    rows = month_recap_with_proof(month)
    
    # --- PERBAIKAN DI SINI ---
    # Kita coba import langsung di sini. Kalau gagal, berarti belum install.
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "Library openpyxl belum terinstall. Jalankan: pip install openpyxl"}), 500
    
    # Kalau berhasil import, lanjut...
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rekap {month}"

    # Header & Judul
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Rekap Presensi Archaic Coffee - {month}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "ID", "ID Pegawai", "Nama", "Jabatan",
        "Tanggal", "Masuk", "Pulang", "Status",
        "Bukti Masuk", "Bukti Pulang",
    ]
    ws.append(headers)
    
    # Styling Header Tabel
    for c in range(1, len(headers) + 1):
        ws.cell(2, c).fill = PatternFill("solid", fgColor="222222")
        ws.cell(2, c).font = Font(bold=True, color="D4AF37")

    # Isi Data
    for r in rows:
        ws.append(list(r))

    # Lebar Kolom
    for col in "ABCDEFGHIJ":
        ws.column_dimensions[col].width = 18

    # Simpan ke Memory (RAM)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"rekap_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export_pdf")
def export_pdf():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    rows = month_recap_with_proof(month)
    if canvas is None:
        return jsonify({"error": "reportlab belum terpasang (pip install reportlab)"}), 500

    path = os.path.join(APP_DIR, "exports", f"rekap_{month}.pdf")
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.pdfgen import canvas as canv
    from reportlab.lib import colors

    c = canv.Canvas(path, pagesize=landscape(A4))
    w, h = landscape(A4)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(w / 2, h - 40, f"Rekap Presensi Archaic Coffee - {month}")
    c.setFont("Helvetica", 9)

    y = h - 70
    headers = ["ID", "ID Pegawai", "Nama", "Jabatan",
               "Tanggal", "Masuk", "Pulang", "Status"]
    widths = [40, 65, 110, 70, 70, 60, 60, 70]
    x0 = 20
    x = x0
    for i, hdr in enumerate(headers):
        c.setFillColor(colors.gold)
        c.rect(x, y, widths[i], 18, fill=1, stroke=0)
        c.setFillColor(colors.black)
        c.drawString(x + 4, y + 5, hdr)
        x += widths[i]
    y -= 20

    c.setFillColor(colors.white)
    for r in rows:
        vals = [
            r[0], r[1], r[2] or "-", r[3] or "-",
            r[4], r[5] or "-", r[6] or "-", r[7] or "-",
        ]
        x = x0
        for i, v in enumerate(vals):
            c.setFillColor(colors.white)
            c.rect(x, y, widths[i], 16, fill=1, stroke=0)
            c.setFillColor(colors.black)
            c.drawString(x + 4, y + 4, str(v))
            x += widths[i]
        y -= 18
        if y < 40:
            c.showPage()
            c.setFont("Helvetica", 9)
            y = h - 40
    c.save()
    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype="application/pdf",
    )

@app.route("/export_employee_excel")
def export_employee_excel():
    emp_id = (request.args.get("id") or "").strip()
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    
    if not emp_id:
        return jsonify({"error": "id pegawai kosong"}), 400
        
    rows = rekap_by_employee(month, emp_id)
    stats = compute_stats(rows)

    # --- PERBAIKAN DI SINI (Sama seperti tadi) ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl belum terpasang (pip install openpyxl)"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = f"{emp_id} {month}"

    ws.merge_cells("A1:H1")
    nama = rows[0][2] if rows else "-"
    jab = rows[0][3] if rows else "-"
    ws["A1"] = f"Rekap Pegawai {emp_id} ({nama}, {jab}) - {month}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([
        "Tanggal", "Masuk", "Pulang", "Status",
        "Menit Telat", "Durasi Kerja (HH:MM:SS)",
        "Bukti In", "Bukti Out",
    ])
    for c in range(1, 8 + 1):
        ws.cell(2, c).fill = PatternFill("solid", fgColor="222222")
        ws.cell(2, c).font = Font(bold=True, color="D4AF37")

    for r in rows:
        tgl, cin, cout, st = r[4], r[5], r[6], r[7]
        late_min = _late_minutes_for(tgl, cin)
        dur = secs_to_hms(_work_secs(cin, cout))
        ws.append([
            tgl,
            cin or "-",
            cout or "-",
            st or "-",
            late_min,
            dur,
            r[8] or "-",
            r[9] or "-",
        ])

    ws.append([])
    ws.append(["RINGKASAN"])
    ws.append(["Hari Hadir", stats["present_days"]])
    ws.append(["Tepat Waktu", stats["on_time"]])
    ws.append(["Terlambat", stats["late"]])
    ws.append(["Total Menit Telat", stats["late_minutes"]])
    ws.append(["Total Jam Kerja", stats["work_hms"]])
    ws.append(["Rata-rata Check-in", stats["avg_checkin"]])
    ws.append(["Rata-rata Check-out", stats["avg_checkout"]])

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 20

    path = os.path.join(APP_DIR, "exports", f"rekap_{emp_id}_{month}.xlsx")
    wb.save(path)
    
    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ========= Download bukti (ZIP) per hari =========
@app.route("/download_day_zip")
def download_day_zip():
    emp_id = (request.args.get("id") or "").strip()
    tgl = (request.args.get("date") or "").strip()
    if not emp_id or not tgl:
        return jsonify({"error": "id dan date wajib"}), 400
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "SELECT bukti_in, bukti_out FROM attendance WHERE id_pegawai=? AND tanggal=?",
        (emp_id, tgl),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Data tidak ditemukan"}), 404
    bi, bo = row
    mem = io.BytesIO()
    with zipfile.ZipFile(
        mem, mode="w", compression=zipfile.ZIP_DEFLATED
    ) as zf:
        if bi and os.path.exists(os.path.join(APP_DIR, bi)):
            zf.write(
                os.path.join(APP_DIR, bi),
                arcname=os.path.basename(bi)
            )
        if bo and os.path.exists(os.path.join(APP_DIR, bo)):
            zf.write(
                os.path.join(APP_DIR, bo),
                arcname=os.path.basename(bo)
            )
    mem.seek(0)
    fname = f"bukti_{emp_id}_{tgl}.zip"
    return send_file(
        mem,
        as_attachment=True,
        download_name=fname,
        mimetype="application/zip",
    )

# ========= Serve bukti foto langsung =========
@app.route("/proofs/<path:filename>")
def serve_proof(filename):
    """
    Melayani file bukti in/out untuk URL seperti:
    /proofs/BRS01_in_20251112_163029.jpg
    """
    return send_from_directory(PROOFS_DIR, filename)

# ========= Izin / Cuti / Terlambat =========
@app.route("/izin/submit", methods=["POST"])
def izin_submit():
    data = request.json or {}
    idp = (data.get("id_pegawai") or "").strip()
    tanggal = (data.get("tanggal") or "").strip()
    jenis = (data.get("jenis") or "").strip()
    alasan = (data.get("alasan") or "").strip() or None

    if not idp or not tanggal or not jenis:
        return jsonify({"error": "id_pegawai, tanggal, dan jenis wajib diisi"}), 400

    try:
        submit_leave(idp, tanggal, jenis, alasan)
        return jsonify({"ok": True, "message": "Pengajuan tersimpan"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/izin/list")
def izin_list():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    rows = list_leave_by_month(month)
    return jsonify(rows)

@app.route("/izin/approve", methods=["POST"])
def izin_approve():
    data = request.json or {}
    id_izin = data.get("id")
    if not id_izin:
        return jsonify({"error": "id izin wajib"}), 400
    try:
        approve_leave(id_izin)
        return jsonify({"ok": True, "message": "Pengajuan disetujui"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/izin/reject", methods=["POST"])
def izin_reject():
    data = request.json or {}
    id_izin = data.get("id")
    alasan_admin = (data.get("alasan_admin") or "").strip() or None
    if not id_izin:
        return jsonify({"error": "id izin wajib"}), 400
    try:
        reject_leave(id_izin, alasan_admin)
        return jsonify({"ok": True, "message": "Pengajuan ditolak"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ========= Leaderboard =========
@app.route("/leaderboard")
def leaderboard():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    rows = leaderboard_month(month)
    return jsonify(rows)

# ========= Kredensial & Backup =========
@app.route("/pegawai/set_login", methods=["POST"])
def api_set_login():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    data = request.json or {}
    idp = (data.get("id_pegawai") or "").strip()
    user = (data.get("username") or "").strip()
    pwd = (data.get("password") or "").strip()
    if not (idp and user and pwd):
        return jsonify({"error": "Lengkapi id_pegawai, username, password"}), 400
    try:
        set_credentials(idp, user, pwd)
        return jsonify({"ok": True, "message": "Kredensial diset"})
    except Exception as e:
        return jsonify({"error": f"Gagal set kredensial: {e}"}), 400

@app.route("/admin/backup_db")
def backup_db():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(APP_DIR, "backups", f"database_{ts}.db")
    shutil.copyfile(DB_PATH, dest)
    return send_file(
        dest,
        as_attachment=True,
        download_name=os.path.basename(dest),
        mimetype="application/octet-stream",
    )

@app.route("/admin/restore_db", methods=["POST"])
def restore_db():
    if not session.get("admin"):
        return jsonify({"error": "Admin only"}), 403
    if "file" not in request.files:
        return jsonify({"error": "Tidak ada file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".db"):
        return jsonify({"error": "File harus .db"}), 400
    tmp = os.path.join(APP_DIR, "backups", "restore_tmp.db")
    f.save(tmp)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = os.path.join(APP_DIR, "backups", f"database_before_restore_{ts}.db")
    shutil.copyfile(DB_PATH, bak)
    shutil.copyfile(tmp, DB_PATH)
    return jsonify({"ok": True, "message": "Restore berhasil. Restart aplikasi disarankan."})

if __name__ == "__main__":
    app.run(debug=True, threaded=True, host='0.0.0.0', port=5000)